"""
سرویس تولید فایل Excel Dynamic با فرمول‌های محاسباتی
این فایل Excel شامل فرمول‌هایی است که محاسبات را در خود Excel انجام می‌دهند
تا کاربر بتواند مسیر محاسبات را ببیند و در صورت نیاز داده‌ها را تغییر دهد
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum, Q
from . import models
from .excel_export import (
    ProjectColors,
    ExcelStyleHelper,
    TableOfContentsSheet,
    ProjectSheet,
    UnitsSheet,
    InvestorsSheet,
    PeriodsSheet,
    InterestRatesSheet,
    TransactionsSheet,
    ExpensesSheet,
    SalesSheet,
    UserProfilesSheet,
)


class NamedRangeHelper:
    """کلاس کمکی برای مدیریت Named Ranges"""
    
    @staticmethod
    def create_named_range(workbook, name, sheet_name, cell_range, comment=''):
        """ایجاد Named Range"""
        try:
            from openpyxl.workbook.defined_name import DefinedName
            
            # حذف Named Range قبلی اگر وجود داشت
            if name in workbook.defined_names:
                del workbook.defined_names[name]
            
            # ایجاد Named Range جدید
            defn = DefinedName(name, attr_text=f"'{sheet_name}'!{cell_range}")
            workbook.defined_names[name] = defn
            
            return True
        except Exception as e:
            print(f"خطا در ایجاد Named Range {name}: {str(e)}")
            return False


class FormulaGuideSheet:
    """شیت راهنمای فرمول‌ها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت راهنمای فرمول‌ها"""
        ws = workbook.create_sheet("📖 راهنمای فرمول‌ها", 1)
        
        # عنوان
        ws['A1'] = 'راهنمای فرمول‌های Excel'
        ws['A1'].font = Font(name='Tahoma', size=16, bold=True, color=ProjectColors.HEADER_BG)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        # بخش 1: Named Ranges
        ws[f'A{row}'] = '📌 Named Ranges (نام‌های تعریف شده)'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True, color=ProjectColors.PROFIT)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SUBHEADER_BG)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # هدر جدول Named Ranges
        headers = ['نام', 'محدوده', 'شرح', 'مثال استفاده']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
        row += 1
        
        # لیست Named Ranges - شیت‌های پایه
        ws[f'A{row}'] = 'شیت‌های پایه:'
        ws[f'A{row}'].font = Font(name='Tahoma', size=11, bold=True, color='000080')
        row += 1
        
        named_ranges_base = [
            ('TransactionAmounts', 'Transactions!$I:$I', 'ستون مبلغ تراکنش‌ها', '=SUM(TransactionAmounts)'),
            ('TransactionTypes', 'Transactions!$J:$J', 'ستون نوع تراکنش', '=COUNTIF(TransactionTypes,"آورده")'),
            ('TransactionInvestors', 'Transactions!$C:$C', 'ستون ID سرمایه‌گذار', '=SUMIF(TransactionInvestors,1,TransactionAmounts)'),
            ('TransactionPeriods', 'Transactions!$E:$E', 'ستون ID دوره', '=SUMIFS(TransactionAmounts,TransactionPeriods,1)'),
            ('ExpenseAmounts', 'Expenses!$F:$F', 'ستون مبلغ هزینه‌ها', '=SUM(ExpenseAmounts)'),
            ('SaleAmounts', 'Sales!$E:$E', 'ستون مبلغ فروش', '=SUM(SaleAmounts)'),
            ('UnitAreas', 'Units!$E:$E', 'ستون متراژ واحدها', '=SUM(UnitAreas)'),
            ('UnitPrices', 'Units!$G:$G', 'ستون قیمت نهایی واحدها', '=SUM(UnitPrices)'),
        ]
        
        for name, range_ref, description, example in named_ranges_base:
            ws.cell(row=row, column=1, value=name).font = Font(name='Courier New', size=10, color='0000FF')
            ws.cell(row=row, column=2, value=range_ref).font = Font(name='Courier New', size=9)
            ws.cell(row=row, column=3, value=description)
            ws.cell(row=row, column=4, value=example).font = Font(name='Courier New', size=9, italic=True)
            row += 1
        
        row += 1
        
        # Named Ranges - شیت محاسباتی
        ws[f'A{row}'] = 'شیت Comprehensive_Metrics:'
        ws[f'A{row}'].font = Font(name='Tahoma', size=11, bold=True, color='000080')
        row += 1
        
        named_ranges_calc = [
            ('TotalUnits', 'Comprehensive_Metrics!$B$4', 'تعداد واحدها', '=TotalUnits'),
            ('TotalArea', 'Comprehensive_Metrics!$B$5', 'متراژ کل', '=TotalArea'),
            ('TotalValue', 'Comprehensive_Metrics!$B$6', 'ارزش کل', '=TotalValue'),
            ('TotalDeposits', 'Comprehensive_Metrics!$B$9', 'آورده کل', '=TotalDeposits'),
            ('TotalWithdrawals', 'Comprehensive_Metrics!$B$10', 'برداشت کل', '=TotalWithdrawals'),
            ('TotalCapital', 'Comprehensive_Metrics!$B$11', 'سرمایه موجود', '=TotalCapital'),
            ('TransactionProfit', 'Comprehensive_Metrics!$B$12', 'سود تراکنش‌ها', '=TransactionProfit'),
            ('TotalProfit', 'Comprehensive_Metrics!$B$12', 'سود کل (alias)', '=TotalProfit'),
            ('TotalBalance', 'Comprehensive_Metrics!$B$13', 'موجودی کل', '=TotalBalance'),
            ('TotalExpenses', 'Comprehensive_Metrics!$B$16', 'هزینه کل', '=TotalExpenses'),
            ('TotalSales', 'Comprehensive_Metrics!$B$17', 'فروش کل', '=TotalSales'),
            ('NetCost', 'Comprehensive_Metrics!$B$18', 'هزینه خالص', '=NetCost'),
            ('BuildingProfit', 'Comprehensive_Metrics!$B$19', 'سود نهایی ساختمان', '=BuildingProfit'),
            ('TransactionProfitPercentage', 'Comprehensive_Metrics!$B$22', 'درصد سود تراکنش‌ها', '=TransactionProfitPercentage'),
            ('BuildingProfitPercentage', 'Comprehensive_Metrics!$B$23', 'درصد سود نهایی ساختمان', '=BuildingProfitPercentage'),
            ('TotalProfitPercentage', 'Comprehensive_Metrics!$B$24', 'درصد سود کل هزینه', '=TotalProfitPercentage'),
            ('AverageConstructionPeriod', 'Comprehensive_Metrics!$B$27', 'دوره متوسط ساخت', '=AverageConstructionPeriod'),
            ('AnnualProfitPercentage', 'Comprehensive_Metrics!$B$28', 'درصد سود سالانه', '=AnnualProfitPercentage'),
            ('MonthlyProfitPercentage', 'Comprehensive_Metrics!$B$29', 'درصد سود ماهانه', '=MonthlyProfitPercentage'),
            ('DailyProfitPercentage', 'Comprehensive_Metrics!$B$30', 'درصد سود روزانه', '=DailyProfitPercentage'),
            ('TotalInvestors', 'Comprehensive_Metrics!$B$33', 'تعداد سرمایه‌گذاران', '=TotalInvestors'),
        ]
        
        for name, range_ref, description, example in named_ranges_calc:
            ws.cell(row=row, column=1, value=name).font = Font(name='Courier New', size=10, color='0000FF')
            ws.cell(row=row, column=2, value=range_ref).font = Font(name='Courier New', size=9)
            ws.cell(row=row, column=3, value=description)
            ws.cell(row=row, column=4, value=example).font = Font(name='Courier New', size=9, italic=True)
            row += 1
        
        row += 2
        
        # بخش 2: فرمول‌های رایج
        ws[f'A{row}'] = '🔢 فرمول‌های رایج'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True, color=ProjectColors.PROFIT)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SUBHEADER_BG)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # هدر جدول فرمول‌ها
        headers2 = ['محاسبه', 'فرمول', 'توضیح']
        for col_num, header in enumerate(headers2, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
        row += 1
        
        # لیست فرمول‌های رایج (شیت ترکیبی جامع)
        formulas = [
            ('آورده کل', '=TotalDeposits', 'جمع تمام تراکنش‌های آورده'),
            ('برداشت کل', '=TotalWithdrawals', 'جمع تمام تراکنش‌های برداشت (منفی)'),
            ('سرمایه موجود', '=TotalCapital', 'آورده منهای برداشت'),
            ('سود تراکنش‌ها', '=TransactionProfit', 'سود سرمایه‌گذاران'),
            ('موجودی کل', '=TotalBalance', 'سرمایه به علاوه سود'),
            ('هزینه کل', '=TotalExpenses', 'جمع تمام هزینه‌ها'),
            ('فروش کل', '=TotalSales', 'جمع تمام فروش‌ها'),
            ('هزینه خالص', '=NetCost', 'هزینه منهای فروش'),
            ('سود نهایی ساختمان', '=BuildingProfit', 'ارزش کل منهای هزینه خالص'),
            ('دوره متوسط ساخت', '=AverageConstructionPeriod', 'میانگین وزنی دوره ساخت'),
            ('درصد سود تراکنش‌ها', '=TransactionProfitPercentage', 'نسبت سود تراکنش‌ها به سرمایه'),
            ('درصد سود نهایی', '=BuildingProfitPercentage', 'نسبت سود نهایی به ارزش کل'),
        ]
        
        for calc, formula, desc in formulas:
            ws.cell(row=row, column=1, value=calc).font = Font(name='Tahoma', size=10, bold=True)
            ws.cell(row=row, column=2, value=formula).font = Font(name='Courier New', size=9, color='006600')
            ws.cell(row=row, column=3, value=desc)
            row += 1
        
        row += 2
        
        # بخش 3: نکات مهم
        ws[f'A{row}'] = '💡 نکات مهم'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True, color=ProjectColors.GOLD)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SUBHEADER_BG)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        notes = [
            '۱. تمام محاسبات با فرمول Excel انجام می‌شوند',
            '۲. می‌توانید داده‌های شیت‌های پایه را ویرایش کنید',
            '۳. شیت‌های محاسباتی به صورت خودکار به‌روز می‌شوند',
            '۴. از Named Ranges برای خوانایی بهتر فرمول‌ها استفاده شده',
            '۵. برای مشاهده فرمول هر سلول، روی آن کلیک کنید',
            '۶. مبالغ برداشت در دیتابیس منفی هستند',
            '۷. برای محاسبات پیچیده، از سلول‌های میانی استفاده شده',
        ]
        
        for note in notes:
            ws.cell(row=row, column=1, value=note)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        
        return ws


class DynamicDashboardSheet:
    """شیت Dashboard با فرمول‌های Excel"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت Dashboard با فرمول"""
        ws = workbook.create_sheet("Dashboard_Dynamic")
        
        # عنوان
        ws['A1'] = f'داشبورد پروژه - {project.name} (با فرمول)'
        ws['A1'].font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # === بخش 1: واحدها ===
        ws[f'A{row}'] = 'واحدها'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # تعداد واحدها
        ws[f'A{row}'] = 'تعداد واحدها'
        ws[f'B{row}'] = '=COUNTA(Units!A:A)-1'
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # متراژ کل
        ws[f'A{row}'] = 'متراژ کل'
        ws[f'B{row}'] = '=SUM(UnitAreas)'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        # ارزش کل
        ws[f'A{row}'] = 'ارزش کل'
        ws[f'B{row}'] = '=SUM(UnitPrices)'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        
        # === بخش 2: اطلاعات مالی ===
        ws[f'A{row}'] = 'اطلاعات مالی'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # آورده کل (اصلاح شده)
        deposits_row = row
        ws[f'A{row}'] = 'آورده کل'
        ws[f'B{row}'] = '=SUMIF(TransactionTypes,"آورده",TransactionAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalDeposits', 'Dashboard_Dynamic', f'$B${row}')
        row += 1
        
        # برداشت کل (اصلاح شده)
        withdrawals_row = row
        ws[f'A{row}'] = 'برداشت کل'
        ws[f'B{row}'] = '=SUMIF(TransactionTypes,"خروج از سرمایه",TransactionAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalWithdrawals', 'Dashboard_Dynamic', f'$B${row}')
        row += 1
        
        # سرمایه موجود
        capital_row = row
        ws[f'A{row}'] = 'سرمایه موجود'
        ws[f'B{row}'] = f'=B{deposits_row}+B{withdrawals_row}'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.CAPITAL)
        NamedRangeHelper.create_named_range(workbook, 'TotalCapital', 'Dashboard_Dynamic', f'$B${row}')
        row += 1
        
        # سود کل (اصلاح شده)
        profit_row = row
        ws[f'A{row}'] = 'سود کل'
        ws[f'B{row}'] = '=SUMIF(TransactionTypes,"سود",TransactionAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.PROFIT)
        NamedRangeHelper.create_named_range(workbook, 'TotalProfit', 'Dashboard_Dynamic', f'$B${row}')
        row += 1
        
        # موجودی کل
        ws[f'A{row}'] = 'موجودی کل'
        ws[f'B{row}'] = f'=B{capital_row}+B{profit_row}'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, size=12)
        row += 2
        
        # === بخش 3: هزینه‌ها ===
        ws[f'A{row}'] = 'هزینه‌ها'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # هزینه کل
        expenses_row = row
        ws[f'A{row}'] = 'هزینه کل'
        ws[f'B{row}'] = '=SUM(ExpenseAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalExpenses', 'Dashboard_Dynamic', f'$B${row}')
        row += 1
        
        # فروش/مرجوعی کل
        sales_row = row
        ws[f'A{row}'] = 'فروش/مرجوعی کل'
        ws[f'B{row}'] = '=SUM(SaleAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalSales', 'Dashboard_Dynamic', f'$B${row}')
        row += 1
        
        # هزینه خالص
        ws[f'A{row}'] = 'هزینه خالص'
        ws[f'B{row}'] = f'=B{expenses_row}-B{sales_row}'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.EXPENSE)
        row += 2
        
        # === بخش 4: سرمایه‌گذاران ===
        ws[f'A{row}'] = 'سرمایه‌گذاران'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # تعداد کل
        ws[f'A{row}'] = 'تعداد کل سرمایه‌گذاران'
        ws[f'B{row}'] = '=COUNTA(Investors!A:A)-1'
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        return ws


class DynamicProfitMetricsSheet:
    """شیت محاسبات سود با فرمول"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت Profit_Metrics با فرمول"""
        ws = workbook.create_sheet("Profit_Metrics_Dynamic")
        
        # عنوان
        ws['A1'] = 'محاسبات سود (با فرمول)'
        ws['A1'].font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # سرمایه کل
        ws[f'A{row}'] = 'سرمایه کل'
        ws[f'B{row}'] = '=TotalCapital'
        ws[f'B{row}'].number_format = '#,##0.00'
        capital_row = row
        row += 1
        
        # سود تراکنش‌ها (سرمایه‌گذاران)
        ws[f'A{row}'] = 'سود تراکنش‌ها (سرمایه‌گذاران)'
        ws[f'B{row}'] = '=TotalProfit'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(color=ProjectColors.PROFIT)
        transaction_profit_row = row
        row += 1
        
        # درصد سود تراکنش‌ها
        ws[f'A{row}'] = 'درصد سود تراکنش‌ها (%)'
        ws[f'B{row}'] = f'=IF(B{capital_row}=0, 0, (B{transaction_profit_row}/B{capital_row})*100)'
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].font = Font(color=ProjectColors.PROFIT)
        transaction_profit_pct_row = row
        row += 2
        
        # سود نهایی ساختمان
        ws[f'A{row}'] = 'سود نهایی ساختمان'
        ws[f'B{row}'] = '=SUM(UnitPrices)-TotalExpenses+TotalSales'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.GOLD)
        building_profit_row = row
        row += 1
        
        # درصد سود نهایی ساختمان
        ws[f'A{row}'] = 'درصد سود نهایی ساختمان (%)'
        ws[f'B{row}'] = '=IF(SUM(UnitPrices)=0,0,((SUM(UnitPrices)-TotalExpenses+TotalSales)/SUM(UnitPrices))*100)'
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.GOLD)
        building_profit_pct_row = row
        row += 2
        
        # محاسبه روزهای فعال (از Project)
        ws[f'A{row}'] = 'روزهای فعال'
        ws[f'B{row}'] = f'=DAYS(Project!F2, Project!D2)'  # تاریخ پایان - تاریخ شروع
        ws[f'B{row}'].number_format = '#,##0'
        active_days_row = row
        row += 1
        
        # درصد سود سالانه (تراکنش‌ها)
        ws[f'A{row}'] = 'درصد سود سالانه تراکنش‌ها (%)'
        ws[f'B{row}'] = f'=IF(B{active_days_row}=0, 0, (B{transaction_profit_pct_row}/B{active_days_row})*365)'
        ws[f'B{row}'].number_format = '0.00'
        row += 1
        
        # درصد سود ماهانه (تراکنش‌ها)
        ws[f'A{row}'] = 'درصد سود ماهانه تراکنش‌ها (%)'
        ws[f'B{row}'] = f'=IF(B{active_days_row}=0, 0, (B{transaction_profit_pct_row}/B{active_days_row})*30)'
        ws[f'B{row}'].number_format = '0.00'
        row += 1
        
        # درصد سود روزانه (تراکنش‌ها)
        ws[f'A{row}'] = 'درصد سود روزانه تراکنش‌ها (%)'
        ws[f'B{row}'] = f'=IF(B{active_days_row}=0, 0, B{transaction_profit_pct_row}/B{active_days_row})'
        ws[f'B{row}'].number_format = '0.0000'
        row += 1
        
        # تنظیم عرض
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        return ws


class DynamicCostMetricsSheet:
    """شیت محاسبات هزینه با فرمول"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت Cost_Metrics با فرمول"""
        ws = workbook.create_sheet("Cost_Metrics_Dynamic")
        
        # عنوان
        ws['A1'] = 'محاسبات هزینه (با فرمول)'
        ws['A1'].font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # هزینه کل
        ws[f'A{row}'] = 'هزینه کل'
        ws[f'B{row}'] = '=TotalExpenses'
        ws[f'B{row}'].number_format = '#,##0.00'
        expenses_row = row
        row += 1
        
        # فروش کل
        ws[f'A{row}'] = 'فروش/مرجوعی کل'
        ws[f'B{row}'] = '=TotalSales'
        ws[f'B{row}'].number_format = '#,##0.00'
        sales_row = row
        row += 1
        
        # هزینه خالص
        ws[f'A{row}'] = 'هزینه خالص'
        ws[f'B{row}'] = f'=B{expenses_row}-B{sales_row}'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.EXPENSE)
        net_cost_row = row
        row += 2
        
        # متراژ کل
        ws[f'A{row}'] = 'متراژ کل'
        ws[f'B{row}'] = '=SUM(UnitAreas)'
        ws[f'B{row}'].number_format = '#,##0.00'
        total_area_row = row
        row += 1
        
        # هزینه هر متر خالص
        ws[f'A{row}'] = 'هزینه هر متر خالص'
        ws[f'B{row}'] = f'=IF(B{total_area_row}=0, 0, B{net_cost_row}/B{total_area_row})'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        # زیربنای کل (از Project)
        ws[f'A{row}'] = 'زیربنای کل'
        ws[f'B{row}'] = '=Project!J2'  # Total_Infrastructure
        ws[f'B{row}'].number_format = '#,##0.00'
        infrastructure_row = row
        row += 1
        
        # هزینه هر متر ناخالص
        ws[f'A{row}'] = 'هزینه هر متر ناخالص'
        ws[f'B{row}'] = f'=IF(B{infrastructure_row}=0, 0, B{net_cost_row}/B{infrastructure_row})'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        
        # ارزش کل
        ws[f'A{row}'] = 'ارزش کل'
        ws[f'B{row}'] = '=SUM(UnitPrices)'
        ws[f'B{row}'].number_format = '#,##0.00'
        total_value_row = row
        row += 1
        
        # ارزش هر متر
        ws[f'A{row}'] = 'ارزش هر متر'
        ws[f'B{row}'] = f'=IF(B{total_area_row}=0, 0, B{total_value_row}/B{total_area_row})'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        
        # سود نهایی
        ws[f'A{row}'] = 'سود نهایی'
        ws[f'B{row}'] = f'=B{total_value_row}-B{net_cost_row}'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.PROFIT)
        final_profit_row = row
        row += 1
        
        # درصد سود کل
        ws[f'A{row}'] = 'درصد سود کل (%)'
        ws[f'B{row}'] = f'=IF(B{net_cost_row}=0, 0, (B{final_profit_row}/B{net_cost_row})*100)'
        ws[f'B{row}'].number_format = '0.00'
        row += 1
        
        # تنظیم عرض
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        return ws


class DynamicInvestorAnalysisSheet:
    """شیت تحلیل سرمایه‌گذاران با فرمول"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت Investor_Analysis با فرمول"""
        ws = workbook.create_sheet("Investor_Analysis_Dynamic")
        
        # هدرها
        headers = [
            'نام سرمایه‌گذار', 'ID', 'آورده کل', 'برداشت کل', 'سرمایه خالص',
            'سود کل', 'موجودی کل', 'نسبت سرمایه (%)', 'نسبت سود (%)', 'شاخص نفع'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
        
        # دریافت همه سرمایه‌گذاران
        investors = models.Investor.objects.all().order_by('last_name', 'first_name')
        
        # نوشتن داده‌ها و فرمول‌ها
        for row_num, investor in enumerate(investors, 2):
            investor_id = investor.id
            investor_name = f"{investor.first_name} {investor.last_name}"
            
            # ستون A: نام (داده خام)
            ws.cell(row=row_num, column=1, value=investor_name)
            
            # ستون B: ID (داده خام)
            ws.cell(row=row_num, column=2, value=investor_id)
            
            # ستون C: آورده کل (فرمول)
            formula_deposits = f'=SUMIFS(TransactionAmounts,TransactionInvestors,B{row_num},TransactionTypes,"آورده")'
            ws.cell(row=row_num, column=3, value=formula_deposits)
            ws.cell(row=row_num, column=3).number_format = '#,##0'
            
            # ستون D: برداشت کل (فرمول)
            formula_withdrawals = f'=SUMIFS(TransactionAmounts,TransactionInvestors,B{row_num},TransactionTypes,"خروج از سرمایه")'
            ws.cell(row=row_num, column=4, value=formula_withdrawals)
            ws.cell(row=row_num, column=4).number_format = '#,##0'
            
            # ستون E: سرمایه خالص (فرمول)
            formula_capital = f'=C{row_num}+D{row_num}'
            ws.cell(row=row_num, column=5, value=formula_capital)
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            
            # ستون F: سود کل (فرمول)
            formula_profit = f'=SUMIFS(TransactionAmounts,TransactionInvestors,B{row_num},TransactionTypes,"سود")'
            ws.cell(row=row_num, column=6, value=formula_profit)
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            
            # ستون G: موجودی کل (فرمول)
            formula_balance = f'=E{row_num}+F{row_num}'
            ws.cell(row=row_num, column=7, value=formula_balance)
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            
            # ستون H: نسبت سرمایه (فرمول)
            formula_capital_ratio = f'=IF(TotalCapital=0,0,(E{row_num}/TotalCapital)*100)'
            ws.cell(row=row_num, column=8, value=formula_capital_ratio)
            ws.cell(row=row_num, column=8).number_format = '0.00'
            
            # ستون I: نسبت سود (فرمول)
            formula_profit_ratio = f'=IF(TotalProfit=0,0,(F{row_num}/TotalProfit)*100)'
            ws.cell(row=row_num, column=9, value=formula_profit_ratio)
            ws.cell(row=row_num, column=9).number_format = '0.00'
            
            # ستون J: شاخص نفع (فرمول)
            formula_index = f'=IF(H{row_num}=0,0,I{row_num}/H{row_num})'
            ws.cell(row=row_num, column=10, value=formula_index)
            ws.cell(row=row_num, column=10).number_format = '0.0000'
            ws.cell(row=row_num, column=10).font = Font(color=ProjectColors.GOLD)
        
        # فریز و فیلتر
        ExcelStyleHelper.freeze_header_row(ws)
        ExcelStyleHelper.add_auto_filter(ws)
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class DynamicPeriodSummarySheet:
    """شیت خلاصه دوره‌ای با فرمول"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت Period_Summary با فرمول"""
        ws = workbook.create_sheet("Period_Summary_Dynamic")
        
        # هدرها
        headers = [
            'دوره', 'سال', 'ماه', 'ID',
            'آورده', 'برداشت', 'سرمایه خالص', 'سود',
            'هزینه', 'فروش', 'مانده صندوق',
            'تجمعی آورده', 'تجمعی برداشت', 'تجمعی سرمایه', 'تجمعی سود'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
        
        # دریافت دوره‌ها
        periods = models.Period.objects.filter(project=project).order_by('year', 'month_number')
        
        # نوشتن داده‌ها و فرمول‌ها
        for row_num, period in enumerate(periods, 2):
            # ستون‌های داده خام
            period_label = f"{period.year}/{period.month_number:02d}"
            ws.cell(row=row_num, column=1, value=period_label)
            ws.cell(row=row_num, column=2, value=period.year)
            ws.cell(row=row_num, column=3, value=period.month_number)
            ws.cell(row=row_num, column=4, value=period.id)
            
            # ستون E: آورده دوره (فرمول)
            formula = f'=SUMIFS(TransactionAmounts,TransactionPeriods,D{row_num},TransactionTypes,"آورده")'
            ws.cell(row=row_num, column=5, value=formula)
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            
            # ستون F: برداشت دوره (فرمول)
            formula = f'=SUMIFS(TransactionAmounts,TransactionPeriods,D{row_num},TransactionTypes,"خروج از سرمایه")'
            ws.cell(row=row_num, column=6, value=formula)
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            
            # ستون G: سرمایه خالص (فرمول)
            formula = f'=E{row_num}+F{row_num}'
            ws.cell(row=row_num, column=7, value=formula)
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            
            # ستون H: سود دوره (فرمول)
            formula = f'=SUMIFS(TransactionAmounts,TransactionPeriods,D{row_num},TransactionTypes,"سود")'
            ws.cell(row=row_num, column=8, value=formula)
            ws.cell(row=row_num, column=8).number_format = '#,##0'
            
            # ستون I: هزینه دوره (فرمول به شیت Expenses - اصلاح شده)
            formula = f'=SUMIF(Expenses!$C:$C,D{row_num},Expenses!$F:$F)'  # ستون C: شناسه دوره، ستون F: مبلغ
            ws.cell(row=row_num, column=9, value=formula)
            ws.cell(row=row_num, column=9).number_format = '#,##0'
            
            # ستون J: فروش دوره (فرمول به شیت Sales - اصلاح شده)
            formula = f'=SUMIF(Sales!$C:$C,D{row_num},Sales!$E:$E)'  # ستون C: شناسه دوره، ستون E: مبلغ
            ws.cell(row=row_num, column=10, value=formula)
            ws.cell(row=row_num, column=10).number_format = '#,##0'
            
            # ستون K: مانده صندوق (فرمول)
            formula = f'=G{row_num}-I{row_num}+J{row_num}'
            ws.cell(row=row_num, column=11, value=formula)
            ws.cell(row=row_num, column=11).number_format = '#,##0'
            
            # ستون L: تجمعی آورده (فرمول)
            if row_num == 2:
                formula = f'=E{row_num}'
            else:
                formula = f'=L{row_num-1}+E{row_num}'
            ws.cell(row=row_num, column=12, value=formula)
            ws.cell(row=row_num, column=12).number_format = '#,##0'
            
            # ستون M: تجمعی برداشت (فرمول)
            if row_num == 2:
                formula = f'=F{row_num}'
            else:
                formula = f'=M{row_num-1}+F{row_num}'
            ws.cell(row=row_num, column=13, value=formula)
            ws.cell(row=row_num, column=13).number_format = '#,##0'
            
            # ستون N: تجمعی سرمایه (فرمول)
            if row_num == 2:
                formula = f'=G{row_num}'
            else:
                formula = f'=N{row_num-1}+G{row_num}'
            ws.cell(row=row_num, column=14, value=formula)
            ws.cell(row=row_num, column=14).number_format = '#,##0'
            
            # ستون O: تجمعی سود (فرمول)
            if row_num == 2:
                formula = f'=H{row_num}'
            else:
                formula = f'=O{row_num-1}+H{row_num}'
            ws.cell(row=row_num, column=15, value=formula)
            ws.cell(row=row_num, column=15).number_format = '#,##0'
        
        # ردیف جمع کل
        total_row = len(periods) + 2
        ws.cell(row=total_row, column=1, value='جمع کل').font = Font(bold=True, size=11)
        
        for col in range(5, 16):  # ستون‌های E تا O
            col_letter = get_column_letter(col)
            formula = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
            cell = ws.cell(row=total_row, column=col, value=formula)
            cell.number_format = '#,##0'
            cell.font = Font(bold=True)
        
        # فریز و فیلتر
        ExcelStyleHelper.freeze_header_row(ws)
        ExcelStyleHelper.add_auto_filter(ws)
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class PeriodExpenseSummarySheet:
    """شیت کمکی برای محاسبه دوره متوسط ساخت"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت خلاصه هزینه‌های دوره‌ای"""
        ws = workbook.create_sheet("PeriodExpenseSummary")
        
        # عنوان
        ws['A1'] = 'خلاصه هزینه‌های دوره‌ای (برای محاسبه دوره متوسط ساخت)'
        ws['A1'].font = Font(name='Tahoma', size=12, bold=True)
        ws.merge_cells('A1:D1')
        
        # هدرها
        headers = ['شناسه دوره', 'عنوان دوره', 'جمع هزینه دوره', 'وزن دوره']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
        
        # دریافت دوره‌ها
        periods = models.Period.objects.filter(project=project).order_by('year', 'month_number')
        
        row = 3
        for period in periods:
            # شناسه دوره
            ws.cell(row=row, column=1, value=period.id)
            
            # عنوان دوره
            ws.cell(row=row, column=2, value=period.label)
            
            # جمع هزینه دوره - با فرمول SUMIF
            ws.cell(row=row, column=3, value=f'=SUMIF(Expenses!$C:$C,A{row},Expenses!$F:$F)')
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            
            # وزن دوره - با فرمول VLOOKUP
            ws.cell(row=row, column=4, value=f'=VLOOKUP(A{row},Periods!$A:$G,7,FALSE)')
            ws.cell(row=row, column=4).number_format = '0.00'
            
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Freeze header
        ws.freeze_panes = 'A3'
        
        return ws


class ComprehensiveMetricsSheet:
    """شیت ترکیبی جامع شامل Dashboard، Profit و Cost Metrics"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت ترکیبی جامع"""
        ws = workbook.create_sheet("Comprehensive_Metrics")
        
        # عنوان اصلی
        ws['A1'] = f'داشبورد جامع - {project.name} (با فرمول)'
        ws['A1'].font = Font(name='Tahoma', size=16, bold=True)
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # ===== بخش واحدها =====
        ws[f'A{row}'] = 'واحدها'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # تعداد واحدها
        ws[f'A{row}'] = 'تعداد واحدها'
        ws[f'B{row}'] = '=COUNTA(Units!A:A)-1'
        ws[f'B{row}'].number_format = '#,##0'
        NamedRangeHelper.create_named_range(workbook, 'TotalUnits', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # متراژ کل
        ws[f'A{row}'] = 'متراژ کل'
        ws[f'B{row}'] = '=SUM(UnitAreas)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalArea', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # ارزش کل
        ws[f'A{row}'] = 'ارزش کل'
        ws[f'B{row}'] = '=SUM(UnitPrices)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalValue', 'Comprehensive_Metrics', f'$B${row}')
        row += 2
        
        # ===== بخش اطلاعات مالی =====
        ws[f'A{row}'] = 'اطلاعات مالی'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # آورده کل
        ws[f'A{row}'] = 'آورده کل'
        ws[f'B{row}'] = '=SUMIF(TransactionTypes,"آورده",TransactionAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalDeposits', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # برداشت کل
        ws[f'A{row}'] = 'برداشت کل'
        ws[f'B{row}'] = '=SUMIF(TransactionTypes,"خروج از سرمایه",TransactionAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalWithdrawals', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # سرمایه موجود
        ws[f'A{row}'] = 'سرمایه موجود'
        ws[f'B{row}'] = f'=TotalDeposits+TotalWithdrawals'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.CAPITAL)
        NamedRangeHelper.create_named_range(workbook, 'TotalCapital', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # سود تراکنش‌ها (سرمایه‌گذاران)
        ws[f'A{row}'] = 'سود تراکنش‌ها (سرمایه‌گذاران)'
        ws[f'B{row}'] = '=SUMIF(TransactionTypes,"سود",TransactionAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(color=ProjectColors.PROFIT)
        NamedRangeHelper.create_named_range(workbook, 'TransactionProfit', 'Comprehensive_Metrics', f'$B${row}')
        NamedRangeHelper.create_named_range(workbook, 'TotalProfit', 'Comprehensive_Metrics', f'$B${row}')  # Alias
        row += 1
        
        # موجودی کل
        ws[f'A{row}'] = 'موجودی کل'
        ws[f'B{row}'] = f'=TotalCapital+TransactionProfit'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalBalance', 'Comprehensive_Metrics', f'$B${row}')
        row += 2
        
        # ===== بخش هزینه‌ها =====
        ws[f'A{row}'] = 'هزینه‌ها'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # هزینه کل
        ws[f'A{row}'] = 'هزینه کل'
        ws[f'B{row}'] = '=SUM(ExpenseAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        NamedRangeHelper.create_named_range(workbook, 'TotalExpenses', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # فروش/مرجوعی کل
        ws[f'A{row}'] = 'فروش/مرجوعی کل'
        ws[f'B{row}'] = '=SUM(SaleAmounts)'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.SALE)
        NamedRangeHelper.create_named_range(workbook, 'TotalSales', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # هزینه خالص
        ws[f'A{row}'] = 'هزینه خالص'
        ws[f'B{row}'] = f'=TotalExpenses-TotalSales'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.EXPENSE)
        NamedRangeHelper.create_named_range(workbook, 'NetCost', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # سود نهایی ساختمان
        ws[f'A{row}'] = 'سود نهایی ساختمان'
        ws[f'B{row}'] = f'=TotalValue-NetCost'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.GOLD)
        NamedRangeHelper.create_named_range(workbook, 'BuildingProfit', 'Comprehensive_Metrics', f'$B${row}')
        row += 2
        
        # ===== بخش محاسبات درصدی =====
        ws[f'A{row}'] = 'محاسبات درصدی'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # درصد سود تراکنش‌ها
        ws[f'A{row}'] = 'درصد سود تراکنش‌ها (%)'
        ws[f'B{row}'] = '=IF(TotalCapital=0,0,(TransactionProfit/TotalCapital)*100)'
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].font = Font(color=ProjectColors.PROFIT)
        NamedRangeHelper.create_named_range(workbook, 'TransactionProfitPercentage', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # درصد سود نهایی ساختمان
        ws[f'A{row}'] = 'درصد سود نهایی ساختمان (%)'
        ws[f'B{row}'] = '=IF(TotalValue=0,0,(BuildingProfit/TotalValue)*100)'
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.GOLD)
        NamedRangeHelper.create_named_range(workbook, 'BuildingProfitPercentage', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # درصد سود کل هزینه
        ws[f'A{row}'] = 'درصد سود کل هزینه (%)'
        ws[f'B{row}'] = '=IF(NetCost=0,0,(BuildingProfit/NetCost)*100)'
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].font = Font(color=ProjectColors.GOLD)
        NamedRangeHelper.create_named_range(workbook, 'TotalProfitPercentage', 'Comprehensive_Metrics', f'$B${row}')
        row += 2
        
        # ===== بخش محاسبات زمانی =====
        ws[f'A{row}'] = 'محاسبات زمانی'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # دوره متوسط ساخت (ماه) - محاسبه پیچیده
        # این محاسبه نیاز به یک شیت کمکی دارد که در PeriodExpenseSummary ایجاد می‌شود
        ws[f'A{row}'] = 'دوره متوسط ساخت (ماه)'
        ws[f'B{row}'] = '=SUMPRODUCT(PeriodExpenseSummary!C:C,PeriodExpenseSummary!D:D)/SUM(PeriodExpenseSummary!C:C)'
        ws[f'B{row}'].number_format = '0.00'
        NamedRangeHelper.create_named_range(workbook, 'AverageConstructionPeriod', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # درصد سود سالانه
        ws[f'A{row}'] = 'درصد سود سالانه (%)'
        ws[f'B{row}'] = '=IF(AverageConstructionPeriod=0,0,(TotalProfitPercentage/AverageConstructionPeriod)*12)'
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.GOLD)
        NamedRangeHelper.create_named_range(workbook, 'AnnualProfitPercentage', 'Comprehensive_Metrics', f'$B${row}')
        annual_profit_row = row
        row += 1
        
        # درصد سود ماهانه
        ws[f'A{row}'] = 'درصد سود ماهانه (%)'
        ws[f'B{row}'] = '=AnnualProfitPercentage/12'
        ws[f'B{row}'].number_format = '0.00'
        NamedRangeHelper.create_named_range(workbook, 'MonthlyProfitPercentage', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # درصد سود روزانه
        ws[f'A{row}'] = 'درصد سود روزانه (%)'
        ws[f'B{row}'] = '=MonthlyProfitPercentage/30'
        ws[f'B{row}'].number_format = '0.0000'
        NamedRangeHelper.create_named_range(workbook, 'DailyProfitPercentage', 'Comprehensive_Metrics', f'$B${row}')
        row += 2
        
        # ===== بخش سرمایه‌گذاران =====
        ws[f'A{row}'] = 'سرمایه‌گذاران'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SECTION_BG)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # تعداد کل سرمایه‌گذاران
        ws[f'A{row}'] = 'تعداد کل سرمایه‌گذاران'
        ws[f'B{row}'] = '=COUNTA(Investors!A:A)-1'
        ws[f'B{row}'].number_format = '#,##0'
        NamedRangeHelper.create_named_range(workbook, 'TotalInvestors', 'Comprehensive_Metrics', f'$B${row}')
        row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        
        return ws


class DynamicTransactionSummarySheet:
    """شیت خلاصه تراکنش‌ها با فرمول"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت Transaction_Summary با فرمول"""
        ws = workbook.create_sheet("Transaction_Summary_Dynamic")
        
        # عنوان
        ws['A1'] = 'خلاصه تراکنش‌ها (با فرمول)'
        ws['A1'].font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # تعداد کل تراکنش‌ها
        ws[f'A{row}'] = 'تعداد کل تراکنش‌ها'
        ws[f'B{row}'] = '=COUNTA(Transactions!A:A)-1'
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # تعداد آورده
        ws[f'A{row}'] = 'تعداد آورده'
        ws[f'B{row}'] = '=COUNTIF(TransactionTypes,"آورده")'
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # تعداد برداشت
        ws[f'A{row}'] = 'تعداد برداشت'
        ws[f'B{row}'] = '=COUNTIF(TransactionTypes,"خروج از سرمایه")'
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # تعداد سود
        ws[f'A{row}'] = 'تعداد سود'
        ws[f'B{row}'] = '=COUNTIF(TransactionTypes,"سود")'
        ws[f'B{row}'].number_format = '#,##0'
        row += 2
        
        # مجموع آورده
        ws[f'A{row}'] = 'مجموع آورده'
        ws[f'B{row}'] = '=TotalDeposits'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        # مجموع برداشت
        ws[f'A{row}'] = 'مجموع برداشت'
        ws[f'B{row}'] = '=TotalWithdrawals'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        # مجموع سود
        ws[f'A{row}'] = 'مجموع سود'
        ws[f'B{row}'] = '=TotalProfit'
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        # سرمایه موجود
        ws[f'A{row}'] = 'سرمایه موجود'
        ws[f'B{row}'] = '=TotalCapital'
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color=ProjectColors.CAPITAL)
        row += 1
        
        # تنظیم عرض
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        return ws


class ExcelDynamicExportService:
    """سرویس اصلی برای تولید Excel با فرمول‌ها"""
    
    def __init__(self, project):
        self.project = project
        self.workbook = Workbook()
    
    def generate_excel(self):
        """تولید فایل Excel کامل با فرمول‌ها"""
        if not self.project:
            raise ValueError('هیچ پروژه‌ای برای export یافت نشد')
        
        # حذف شیت پیش‌فرض
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # ایجاد شیت فهرست
        TableOfContentsSheet.create(self.workbook, self.project)
        
        # ایجاد شیت راهنما
        FormulaGuideSheet.create(self.workbook, self.project)
        
        # ایجاد شیت‌های داده پایه
        self.create_base_sheets()
        
        # ایجاد شیت کمکی برای محاسبات
        PeriodExpenseSummarySheet.create(self.workbook, self.project)
        
        # ایجاد شیت ترکیبی جامع
        ComprehensiveMetricsSheet.create(self.workbook, self.project)
        
        # سایر شیت‌های محاسباتی
        DynamicInvestorAnalysisSheet.create(self.workbook, self.project)
        DynamicPeriodSummarySheet.create(self.workbook, self.project)
        DynamicTransactionSummarySheet.create(self.workbook, self.project)
        
        # در آخر Named Ranges برای شیت‌های پایه
        self.create_named_ranges()
        
        return self.workbook
    
    def create_base_sheets(self):
        """ایجاد شیت‌های داده پایه (بدون فرمول)"""
        # استفاده از کلاس‌های موجود
        ProjectSheet.create(self.workbook, self.project)
        UnitsSheet.create(self.workbook, self.project)
        InvestorsSheet.create(self.workbook, self.project)
        PeriodsSheet.create(self.workbook, self.project)
        InterestRatesSheet.create(self.workbook, self.project)
        TransactionsSheet.create(self.workbook, self.project)
        ExpensesSheet.create(self.workbook, self.project)
        SalesSheet.create(self.workbook, self.project)
        UserProfilesSheet.create(self.workbook, self.project)
        
        # فریز و فیلتر
        base_sheets = ['Units', 'Investors', 'Periods', 'InterestRates', 
                      'Transactions', 'Expenses', 'Sales', 'UserProfiles']
        for sheet_name in base_sheets:
            if sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
                ExcelStyleHelper.freeze_header_row(ws)
                ExcelStyleHelper.add_auto_filter(ws)
    
    def create_named_ranges(self):
        """ایجاد Named Ranges برای ستون‌های مهم"""
        # Transactions (اصلاح شده)
        NamedRangeHelper.create_named_range(
            self.workbook, 'TransactionAmounts', 'Transactions', '$I:$I'  # ستون 9: مبلغ
        )
        NamedRangeHelper.create_named_range(
            self.workbook, 'TransactionTypes', 'Transactions', '$J:$J'   # ستون 10: نوع تراکنش
        )
        NamedRangeHelper.create_named_range(
            self.workbook, 'TransactionInvestors', 'Transactions', '$C:$C'  # ستون 3: شناسه سرمایه‌گذار
        )
        NamedRangeHelper.create_named_range(
            self.workbook, 'TransactionPeriods', 'Transactions', '$E:$E'    # ستون 5: شناسه دوره
        )
        
        # Expenses (اصلاح شده)
        NamedRangeHelper.create_named_range(
            self.workbook, 'ExpenseAmounts', 'Expenses', '$F:$F'  # ستون 6: مبلغ
        )
        
        # Sales (اصلاح شده)
        NamedRangeHelper.create_named_range(
            self.workbook, 'SaleAmounts', 'Sales', '$E:$E'  # ستون 5: مبلغ
        )
        
        # Units (اصلاح شده)
        NamedRangeHelper.create_named_range(
            self.workbook, 'UnitAreas', 'Units', '$E:$E'    # ستون 5: متراژ
        )
        NamedRangeHelper.create_named_range(
            self.workbook, 'UnitPrices', 'Units', '$G:$G'   # ستون 7: قیمت نهایی
        )
    

