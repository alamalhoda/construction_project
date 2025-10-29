"""
سرویس تولید فایل Excel برای export کامل اطلاعات و محاسبات پروژه
این فایل شامل کلاس‌های مختلفی است که هر کدام یک شیت از فایل Excel را تولید می‌کنند
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum, Q
from . import models
from .calculations import (
    ProjectCalculations,
    ProfitCalculations,
    InvestorCalculations,
    TransactionCalculations
)


class ProjectColors:
    """رنگ‌های استاندارد پروژه"""
    # رنگ‌های اصلی
    DEPOSIT = '2185d0'           # آبی - آورده و واریزی
    WITHDRAWAL = 'db2828'        # قرمز - برداشت و خروجی
    PROFIT = '21ba45'            # سبز - سود مشارکت و درآمد
    CAPITAL = 'aa26ff'           # بنفش - سرمایه موجود و موجودی
    EXPENSE = 'dc3545'           # قرمز تیره - هزینه‌ها و خرجی
    SALE = 'ffc107'              # زرد - فروش/مرجوعی
    BALANCE = '6c757d'           # خاکستری - مانده صندوق و مجموع
    GOLD = 'ffd700'              # طلایی - شاخص نفع و عملکرد برتر
    
    # رنگ‌های کمکی
    HEADER_BG = '4472C4'         # آبی تیره - هدر اصلی
    SUBHEADER_BG = 'B4C7E7'      # آبی روشن - هدر فرعی
    SECTION_BG = 'E7E6E6'        # خاکستری روشن - بخش‌ها
    WHITE = 'FFFFFF'             # سفید


class ExcelStyleHelper:
    """کلاس کمکی برای استایل‌های Excel"""
    
    @staticmethod
    def normalize_datetime(dt):
        """حذف timezone از datetime برای سازگاری با Excel"""
        if dt is None:
            return None
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            # تبدیل به naive datetime (بدون timezone)
            return dt.replace(tzinfo=None)
        return dt
    
    @staticmethod
    def get_header_font():
        """فونت هدر"""
        return Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
    
    @staticmethod
    def get_header_fill(color=None):
        """پس‌زمینه هدر با رنگ قابل تنظیم"""
        if color is None:
            color = ProjectColors.HEADER_BG
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    @staticmethod
    def get_colored_fill(color):
        """پس‌زمینه با رنگ مشخص"""
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    @staticmethod
    def get_default_font():
        """فونت پیش‌فرض"""
        return Font(name='Tahoma', size=10)
    
    @staticmethod
    def get_border():
        """حاشیه سلول"""
        thin_border = Side(border_style='thin', color='000000')
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    @staticmethod
    def get_center_alignment():
        """تراز وسط"""
        return Alignment(horizontal='center', vertical='center')
    
    @staticmethod
    def apply_header_style(cell):
        """اعمال استایل به هدر"""
        cell.font = ExcelStyleHelper.get_header_font()
        cell.fill = ExcelStyleHelper.get_header_fill()
        cell.alignment = ExcelStyleHelper.get_center_alignment()
        cell.border = ExcelStyleHelper.get_border()
    
    @staticmethod
    def apply_cell_style(cell):
        """اعمال استایل به سلول عادی"""
        cell.font = ExcelStyleHelper.get_default_font()
        cell.border = ExcelStyleHelper.get_border()
    
    @staticmethod
    def auto_adjust_column_width(worksheet):
        """تنظیم خودکار عرض ستون‌ها"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # حداکثر 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def freeze_header_row(worksheet):
        """فریز کردن ردیف اول (هدر)"""
        worksheet.freeze_panes = 'A2'
    
    @staticmethod
    def add_auto_filter(worksheet):
        """اضافه کردن فیلتر خودکار"""
        if worksheet.max_row > 1:  # فقط اگر داده وجود داشته باشد
            worksheet.auto_filter.ref = worksheet.dimensions


class ChartsSheet:
    """شیت نمودارها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت با نمودارهای تحلیلی"""
        ws = workbook.create_sheet("📊 نمودارها", 1)  # بعد از فهرست
        
        # عنوان
        ws['A1'] = f'نمودارهای تحلیلی - {project.name}'
        ws['A1'].font = Font(name='Tahoma', size=16, bold=True, color=ProjectColors.HEADER_BG)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # دریافت داده‌ها
        stats = ProjectCalculations.calculate_project_statistics(project.id)
        
        row = 3
        
        # === نمودار 1: نمودار دایره‌ای توزیع سرمایه بین سرمایه‌گذاران ===
        ws[f'A{row}'] = '📊 توزیع سرمایه بین سرمایه‌گذاران (Top 10)'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        row += 1
        
        # هدر داده‌ها
        ws[f'A{row}'] = 'سرمایه‌گذار'
        ws[f'B{row}'] = 'سرمایه (تومان)'
        ws[f'A{row}'].font = ExcelStyleHelper.get_header_font()
        ws[f'B{row}'].font = ExcelStyleHelper.get_header_font()
        ws[f'A{row}'].fill = ExcelStyleHelper.get_header_fill()
        ws[f'B{row}'].fill = ExcelStyleHelper.get_header_fill()
        data_start_row = row + 1
        row += 1
        
        # دریافت برترین سرمایه‌گذاران
        top_investors = models.Investor.objects.annotate(
            total_principal=Sum('transaction__amount', filter=Q(transaction__transaction_type__in=['principal_deposit','loan_deposit']))
        ).filter(total_principal__isnull=False).order_by('-total_principal')[:10]
        
        for investor in top_investors:
            ws[f'A{row}'] = f'{investor.first_name} {investor.last_name}'
            ws[f'B{row}'] = float(investor.total_principal or 0)
            ws[f'B{row}'].number_format = '#,##0'
            row += 1
        
        data_end_row = row - 1
        
        # ایجاد نمودار دایره‌ای
        pie_chart = PieChart()
        pie_chart.title = "توزیع سرمایه بین سرمایه‌گذاران"
        pie_chart.height = 12
        pie_chart.width = 18
        
        labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        data = Reference(ws, min_col=2, min_row=data_start_row-1, max_row=data_end_row)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        # تنظیمات نمایش
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        
        ws.add_chart(pie_chart, f'D{data_start_row-1}')
        
        row += 3
        
        # === نمودار 2: نمودار میله‌ای مقایسه آورده و برداشت ===
        ws[f'A{row}'] = '📊 مقایسه آورده و برداشت سرمایه‌گذاران (Top 10)'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        row += 1
        
        # هدر داده‌ها
        ws[f'A{row}'] = 'سرمایه‌گذار'
        ws[f'B{row}'] = 'آورده'
        ws[f'C{row}'] = 'برداشت'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = ExcelStyleHelper.get_header_font()
            ws[f'{col}{row}'].fill = ExcelStyleHelper.get_header_fill()
        bar_data_start = row + 1
        row += 1
        
        # داده‌ها
        top_investors_trans = models.Investor.objects.annotate(
            total_deposits=Sum('transaction__amount', filter=Q(transaction__transaction_type__in=['principal_deposit','loan_deposit'])),
            total_withdrawals=Sum('transaction__amount', filter=Q(transaction__transaction_type='principal_withdrawal'))
        ).filter(total_deposits__isnull=False).order_by('-total_deposits')[:10]
        
        for investor in top_investors_trans:
            ws[f'A{row}'] = f'{investor.first_name} {investor.last_name}'
            ws[f'B{row}'] = float(investor.total_deposits or 0)
            ws[f'C{row}'] = abs(float(investor.total_withdrawals or 0))
            ws[f'B{row}'].number_format = '#,##0'
            ws[f'C{row}'].number_format = '#,##0'
            row += 1
        
        bar_data_end = row - 1
        
        # ایجاد نمودار میله‌ای
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "مقایسه آورده و برداشت"
        bar_chart.height = 12
        bar_chart.width = 18
        bar_chart.y_axis.title = 'مبلغ (تومان)'
        
        data = Reference(ws, min_col=2, min_row=bar_data_start-1, max_row=bar_data_end, max_col=3)
        cats = Reference(ws, min_col=1, min_row=bar_data_start, max_row=bar_data_end)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        ws.add_chart(bar_chart, f'E{bar_data_start-1}')
        
        row += 3
        
        # === نمودار 3: نمودار خطی روند دوره‌ای ===
        ws[f'A{row}'] = '📊 روند دوره‌ای (سرمایه، هزینه، فروش)'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True)
        row += 1
        
        # هدر داده‌ها
        ws[f'A{row}'] = 'دوره'
        ws[f'B{row}'] = 'سرمایه'
        ws[f'C{row}'] = 'هزینه'
        ws[f'D{row}'] = 'فروش'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = ExcelStyleHelper.get_header_font()
            ws[f'{col}{row}'].fill = ExcelStyleHelper.get_header_fill()
        line_data_start = row + 1
        row += 1
        
        # دریافت داده‌های دوره‌ای
        from construction.api import PeriodViewSet
        period_data = []
        try:
            # استفاده از API موجود
            periods = models.Period.objects.filter(project=project).order_by('year', 'month_number')
            
            cumulative_capital = 0
            cumulative_expenses = 0
            cumulative_sales = 0
            
            for period in periods[:12]:  # فقط 12 دوره اول
                # محاسبه سرمایه دوره
                deposits = models.Transaction.objects.filter(
                    period=period, 
                    transaction_type='principal_deposit'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                withdrawals = models.Transaction.objects.filter(
                    period=period,
                    transaction_type='principal_withdrawal'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                period_capital = float(deposits) + float(withdrawals)
                cumulative_capital += period_capital
                
                # هزینه‌های دوره
                period_expenses = models.Expense.objects.filter(
                    period=period
                ).aggregate(total=Sum('amount'))['total'] or 0
                cumulative_expenses += float(period_expenses)
                
                # فروش دوره
                period_sales = models.Sale.objects.filter(
                    period=period
                ).aggregate(total=Sum('amount'))['total'] or 0
                cumulative_sales += float(period_sales)
                
                period_data.append({
                    'label': f'{period.year}/{period.month_number}',
                    'capital': cumulative_capital / 1000000,  # به میلیون تومان
                    'expenses': cumulative_expenses / 1000000,
                    'sales': cumulative_sales / 1000000
                })
        except:
            pass
        
        # نوشتن داده‌ها
        for pd in period_data:
            ws[f'A{row}'] = pd['label']
            ws[f'B{row}'] = pd['capital']
            ws[f'C{row}'] = pd['expenses']
            ws[f'D{row}'] = pd['sales']
            for col in ['B', 'C', 'D']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            row += 1
        
        line_data_end = row - 1
        
        if period_data:
            # ایجاد نمودار خطی
            line_chart = LineChart()
            line_chart.title = "روند تجمعی دوره‌ای (میلیون تومان)"
            line_chart.height = 12
            line_chart.width = 18
            line_chart.y_axis.title = 'مبلغ (میلیون تومان)'
            line_chart.x_axis.title = 'دوره'
            
            data = Reference(ws, min_col=2, min_row=line_data_start-1, max_row=line_data_end, max_col=4)
            cats = Reference(ws, min_col=1, min_row=line_data_start, max_row=line_data_end)
            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(cats)
            
            ws.add_chart(line_chart, f'F{line_data_start-1}')
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        return ws


class ExecutiveSummarySheet:
    """شیت خلاصه اجرایی"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت خلاصه اجرایی با KPI های کلیدی"""
        ws = workbook.create_sheet("📊 خلاصه اجرایی", 1)  # بعد از فهرست
        
        # دریافت محاسبات
        stats = ProjectCalculations.calculate_project_statistics(project.id)
        profit_metrics = ProfitCalculations.calculate_profit_percentages(project.id)
        
        # عنوان اصلی
        ws['A1'] = f'خلاصه اجرایی - {project.name}'
        ws['A1'].font = Font(name='Tahoma', size=18, bold=True, color=ProjectColors.CAPITAL)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # تاریخ
        ws['A2'] = f'تاریخ گزارش: {timezone.now().strftime("%Y/%m/%d - %H:%M")}'
        ws['A2'].font = Font(name='Tahoma', size=10, italic=True)
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        row = 4
        
        # === بخش 1: KPI های کلیدی ===
        ws[f'A{row}'] = '📈 شاخص‌های کلیدی عملکرد (KPI)'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True, color=ProjectColors.HEADER_BG)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SUBHEADER_BG)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # KPI ها در قالب کارت
        kpis = [
            ('💰 سرمایه موجود', stats.get('transaction_statistics', {}).get('net_principal', 0), 'تومان', ProjectColors.CAPITAL),
            ('💵 سود کل', stats.get('transaction_statistics', {}).get('total_profits', 0), 'تومان', ProjectColors.PROFIT),
            ('💸 هزینه خالص', stats.get('cost_metrics', {}).get('net_cost', 0), 'تومان', ProjectColors.EXPENSE),
            ('🏠 تعداد واحدها', stats.get('units_statistics', {}).get('total_units', 0), 'واحد', ProjectColors.BALANCE),
            ('👥 سرمایه‌گذاران', stats.get('investor_statistics', {}).get('total_investors', 0), 'نفر', ProjectColors.BALANCE),
            ('📊 درصد سود کل', profit_metrics.get('total_profit_percentage', 0), '%', ProjectColors.GOLD),
        ]
        
        col = 0
        for label, value, unit, color in kpis:
            if col >= 3:
                row += 2
                col = 0
            
            cell_col = chr(65 + col * 2)  # A, C, E
            ws[f'{cell_col}{row}'] = label
            ws[f'{cell_col}{row}'].font = Font(name='Tahoma', size=10, bold=True)
            ws[f'{cell_col}{row}'].fill = ExcelStyleHelper.get_colored_fill(color)
            ws[f'{cell_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'{cell_col}{row+1}'] = value
            ws[f'{cell_col}{row+1}'].font = Font(name='Tahoma', size=14, bold=True, color=color)
            ws[f'{cell_col}{row+1}'].alignment = Alignment(horizontal='center', vertical='center')
            if isinstance(value, (int, float)) and unit == 'تومان':
                ws[f'{cell_col}{row+1}'].number_format = '#,##0'
            elif unit == '%':
                ws[f'{cell_col}{row+1}'].number_format = '0.00'
            
            next_col = chr(65 + col * 2 + 1)
            ws.merge_cells(f'{cell_col}{row}:{next_col}{row}')
            ws.merge_cells(f'{cell_col}{row+1}:{next_col}{row+1}')
            
            col += 1
        
        row += 3
        
        # === بخش 2: خلاصه مالی ===
        ws[f'A{row}'] = '💰 خلاصه مالی'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True, color=ProjectColors.HEADER_BG)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SUBHEADER_BG)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # هدر جدول مالی
        financial_headers = ['شرح', 'مبلغ (تومان)', 'درصد']
        for col_num, header in enumerate(financial_headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
            cell.alignment = ExcelStyleHelper.get_center_alignment()
        row += 1
        
        # داده‌های مالی
        financial_data = [
            ('آورده کل', stats.get('transaction_statistics', {}).get('total_deposits', 0), ''),
            ('برداشت کل', stats.get('transaction_statistics', {}).get('total_withdrawals', 0), ''),
            ('سرمایه خالص', stats.get('transaction_statistics', {}).get('net_principal', 0), '100%'),
            ('سود کل', stats.get('transaction_statistics', {}).get('total_profits', 0), ''),
            ('موجودی کل', stats.get('transaction_statistics', {}).get('grand_total', 0), ''),
            ('', '', ''),  # خط جداکننده
            ('هزینه‌ها', stats.get('cost_metrics', {}).get('total_expenses', 0), ''),
            ('فروش/مرجوعی', stats.get('cost_metrics', {}).get('total_sales', 0), ''),
            ('هزینه خالص', stats.get('cost_metrics', {}).get('net_cost', 0), ''),
        ]
        
        for desc, amount, pct in financial_data:
            ws.cell(row=row, column=1, value=desc).font = Font(name='Tahoma', size=10, bold=(desc == ''))
            if isinstance(amount, (int, float)):
                cell = ws.cell(row=row, column=2, value=float(amount))
                cell.number_format = '#,##0'
            ws.cell(row=row, column=3, value=pct)
            row += 1
        
        row += 1
        
        # === بخش 3: برترین سرمایه‌گذاران ===
        ws[f'A{row}'] = '🏆 برترین سرمایه‌گذاران (بر اساس سرمایه)'
        ws[f'A{row}'].font = Font(name='Tahoma', size=14, bold=True, color=ProjectColors.HEADER_BG)
        ws[f'A{row}'].fill = ExcelStyleHelper.get_colored_fill(ProjectColors.SUBHEADER_BG)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # هدر جدول سرمایه‌گذاران
        investor_headers = ['رتبه', 'نام', 'سرمایه (تومان)', 'سود (تومان)', 'درصد مالکیت']
        for col_num, header in enumerate(investor_headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = ExcelStyleHelper.get_header_font()
            cell.fill = ExcelStyleHelper.get_header_fill()
            cell.alignment = ExcelStyleHelper.get_center_alignment()
        row += 1
        
        # دریافت برترین سرمایه‌گذاران
        from django.db.models import Sum
        top_investors = models.Investor.objects.annotate(
            total_principal=Sum('transaction__amount', filter=Q(transaction__transaction_type__in=['principal_deposit','loan_deposit'])),
            total_profit=Sum('transaction__amount', filter=Q(transaction__transaction_type='profit_accrual'))
        ).filter(total_principal__isnull=False).order_by('-total_principal')[:5]
        
        for rank, investor in enumerate(top_investors, 1):
            ws.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=f'{investor.first_name} {investor.last_name}')
            
            cell = ws.cell(row=row, column=3, value=float(investor.total_principal or 0))
            cell.number_format = '#,##0'
            
            cell = ws.cell(row=row, column=4, value=float(investor.total_profit or 0))
            cell.number_format = '#,##0'
            
            # محاسبه درصد مالکیت
            try:
                ownership = InvestorCalculations.calculate_investor_ownership(investor.id, project.id)
                ownership_pct = ownership.get('ownership_percentage', 0)
            except:
                ownership_pct = 0
            
            cell = ws.cell(row=row, column=5, value=ownership_pct)
            cell.number_format = '0.00'
            
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        return ws


class TableOfContentsSheet:
    """شیت فهرست محتوا"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت فهرست محتوا"""
        ws = workbook.create_sheet("📋 فهرست", 0)  # در ابتدا قرار می‌گیرد
        
        # عنوان اصلی
        ws['A1'] = f'گزارش جامع پروژه: {project.name}'
        ws['A1'].font = Font(name='Tahoma', size=16, bold=True, color=ProjectColors.CAPITAL)
        ws.merge_cells('A1:C1')
        
        # زمان تولید
        ws['A2'] = f'تاریخ تولید: {timezone.now().strftime("%Y/%m/%d - %H:%M:%S")}'
        ws['A2'].font = Font(name='Tahoma', size=10, italic=True)
        ws.merge_cells('A2:C2')
        
        row = 4
        
        # بخش داده‌های پایه
        ws[f'A{row}'] = '📊 داده‌های پایه'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color=ProjectColors.HEADER_BG)
        row += 1
        
        base_sheets = [
            ('Project', 'اطلاعات پروژه', '📁'),
            ('Units', 'واحدها', '🏠'),
            ('Investors', 'سرمایه‌گذاران', '👥'),
            ('Periods', 'دوره‌ها', '📅'),
            ('InterestRates', 'نرخ‌های سود', '💹'),
            ('Transactions', 'تراکنش‌ها', '💳'),
            ('Expenses', 'هزینه‌ها', '💰'),
            ('Sales', 'فروش/مرجوعی', '🏷️'),
            ('UserProfiles', 'پروفایل کاربران', '👤'),
        ]
        
        for sheet_name, persian_name, icon in base_sheets:
            ws[f'A{row}'] = f'{icon} {persian_name}'
            ws[f'A{row}'].font = Font(name='Tahoma', size=10, color='0000FF', underline='single')
            ws[f'A{row}'].hyperlink = f'#{sheet_name}!A1'
            row += 1
        
        row += 1
        
        # بخش محاسبات
        ws[f'A{row}'] = '📈 محاسبات و تحلیل‌ها'
        ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color=ProjectColors.PROFIT)
        row += 1
        
        calc_sheets = [
            ('Dashboard', 'داشبورد', '📊'),
            ('Profit_Metrics', 'محاسبات سود', '💵'),
            ('Cost_Metrics', 'محاسبات هزینه', '💸'),
            ('Investor_Analysis', 'تحلیل سرمایه‌گذاران', '👥'),
            ('Period_Summary', 'خلاصه دوره‌ای', '📅'),
            ('Transaction_Summary', 'خلاصه تراکنش‌ها', '💳'),
        ]
        
        for sheet_name, persian_name, icon in calc_sheets:
            ws[f'A{row}'] = f'{icon} {persian_name}'
            ws[f'A{row}'].font = Font(name='Tahoma', size=10, color='0000FF', underline='single')
            ws[f'A{row}'].hyperlink = f'#{sheet_name}!A1'
            row += 1
        
        # تنظیم عرض ستون
        ws.column_dimensions['A'].width = 40
        
        return ws


class ProjectSheet:
    """شیت اطلاعات پروژه"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت اطلاعات پروژه"""
        ws = workbook.create_sheet(title="Project")
        
        # هدرها
        headers = [
            'ID', 'نام پروژه', 'تاریخ شروع (شمسی)', 'تاریخ شروع (میلادی)',
            'تاریخ پایان (شمسی)', 'تاریخ پایان (میلادی)', 'فعال',
            'زیربنای کل', 'ضریب اصلاحی', 'تاریخ ایجاد', 'تاریخ به‌روزرسانی'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # نوشتن داده‌ها
        data = [
            project.id,
            project.name,
            str(project.start_date_shamsi),
            project.start_date_gregorian,
            str(project.end_date_shamsi),
            project.end_date_gregorian,
            'بله' if project.is_active else 'خیر',
            float(project.total_infrastructure),
            float(project.correction_factor),
            ExcelStyleHelper.normalize_datetime(project.created_at),
            ExcelStyleHelper.normalize_datetime(project.updated_at)
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            ExcelStyleHelper.apply_cell_style(cell)
            
            # فرمت اعداد
            if col_num == 8:  # زیربنای کل
                cell.number_format = '#,##0.00'
            elif col_num == 9:  # ضریب اصلاحی
                cell.number_format = '0.0000000000'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class UnitsSheet:
    """شیت واحدها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت واحدها"""
        ws = workbook.create_sheet(title="Units")
        
        # هدرها
        headers = [
            'ID', 'شناسه پروژه', 'نام پروژه', 'نام واحد',
            'متراژ', 'قیمت هر متر', 'قیمت نهایی', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت واحدها
        units = models.Unit.objects.filter(project=project).order_by('id')
        
        # نوشتن داده‌ها
        for row_num, unit in enumerate(units, 2):
            data = [
                unit.id,
                unit.project.id,
                unit.project.name,
                unit.name,
                float(unit.area),
                float(unit.price_per_meter),
                float(unit.total_price),
                ExcelStyleHelper.normalize_datetime(unit.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت اعداد
                if col_num in [5, 6, 7]:  # متراژ، قیمت
                    cell.number_format = '#,##0.00'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class InvestorsSheet:
    """شیت سرمایه‌گذاران"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت سرمایه‌گذاران"""
        ws = workbook.create_sheet(title="Investors")
        
        # هدرها
        headers = [
            'ID', 'نام', 'نام خانوادگی', 'شماره تماس', 'ایمیل',
            'نوع مشارکت', 'واحدها', 'تاریخ قرارداد (شمسی)', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت همه سرمایه‌گذاران
        investors = models.Investor.objects.all().order_by('id')
        
        # نوشتن داده‌ها
        for row_num, investor in enumerate(investors, 2):
            # دریافت واحدهای این سرمایه‌گذار در این پروژه
            investor_units = investor.units.filter(project=project)
            units_str = ', '.join([unit.name for unit in investor_units])
            
            data = [
                investor.id,
                investor.first_name,
                investor.last_name,
                investor.phone,
                investor.email or '',
                investor.get_participation_type_display(),
                units_str,
                str(investor.contract_date_shamsi) if investor.contract_date_shamsi else '',
                ExcelStyleHelper.normalize_datetime(investor.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class PeriodsSheet:
    """شیت دوره‌ها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت دوره‌ها"""
        ws = workbook.create_sheet(title="Periods")
        
        # هدرها
        headers = [
            'ID', 'شناسه پروژه', 'عنوان دوره', 'سال شمسی', 'شماره ماه',
            'نام ماه', 'وزن دوره', 'تاریخ شروع (شمسی)', 'تاریخ شروع (میلادی)',
            'تاریخ پایان (شمسی)', 'تاریخ پایان (میلادی)'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت دوره‌ها
        periods = models.Period.objects.filter(project=project).order_by('year', 'month_number')
        
        # نوشتن داده‌ها
        for row_num, period in enumerate(periods, 2):
            data = [
                period.id,
                period.project.id,
                period.label,
                period.year,
                period.month_number,
                period.month_name,
                period.weight,
                str(period.start_date_shamsi),
                period.start_date_gregorian,
                str(period.end_date_shamsi),
                period.end_date_gregorian
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class InterestRatesSheet:
    """شیت نرخ‌های سود"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت نرخ‌های سود"""
        ws = workbook.create_sheet(title="InterestRates")
        
        # هدرها
        headers = [
            'ID', 'نرخ سود روزانه', 'تاریخ اعمال (شمسی)',
            'تاریخ اعمال (میلادی)', 'توضیحات', 'فعال', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت نرخ‌های سود (همه نرخ‌ها، نه فقط مربوط به یک پروژه خاص)
        interest_rates = models.InterestRate.objects.all().order_by('-effective_date')
        
        # نوشتن داده‌ها
        for row_num, rate in enumerate(interest_rates, 2):
            data = [
                rate.id,
                float(rate.rate),
                str(rate.effective_date),
                rate.effective_date_gregorian,
                rate.description,
                'بله' if rate.is_active else 'خیر',
                ExcelStyleHelper.normalize_datetime(rate.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت نرخ سود
                if col_num == 2:
                    cell.number_format = '0.000000000000000'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class TransactionsSheet:
    """شیت تراکنش‌ها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت تراکنش‌ها"""
        ws = workbook.create_sheet(title="Transactions")
        
        # هدرها
        headers = [
            'ID', 'شناسه پروژه', 'شناسه سرمایه‌گذار', 'نام سرمایه‌گذار',
            'شناسه دوره', 'عنوان دوره', 'تاریخ (شمسی)', 'تاریخ (میلادی)',
            'مبلغ', 'نوع تراکنش', 'توضیحات', 'روز مانده', 'روز از شروع',
            'شناسه نرخ سود', 'تولید سیستمی', 'شناسه تراکنش اصلی', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت تراکنش‌ها
        transactions = models.Transaction.objects.filter(project=project).select_related(
            'investor', 'period', 'interest_rate', 'parent_transaction'
        ).order_by('date_gregorian', 'id')
        
        # نوشتن داده‌ها
        for row_num, trans in enumerate(transactions, 2):
            data = [
                trans.id,
                trans.project.id,
                trans.investor.id,
                f"{trans.investor.first_name} {trans.investor.last_name}",
                trans.period.id,
                trans.period.label,
                str(trans.date_shamsi),
                trans.date_gregorian,
                float(trans.amount),
                trans.get_transaction_type_display(),
                trans.description,
                trans.day_remaining,
                trans.day_from_start,
                trans.interest_rate.id if trans.interest_rate else '',
                'بله' if trans.is_system_generated else 'خیر',
                trans.parent_transaction.id if trans.parent_transaction else '',
                ExcelStyleHelper.normalize_datetime(trans.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت مبلغ
                if col_num == 9:
                    cell.number_format = '#,##0.00'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class ExpensesSheet:
    """شیت هزینه‌ها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت هزینه‌ها"""
        ws = workbook.create_sheet(title="Expenses")
        
        # هدرها
        headers = [
            'ID', 'شناسه پروژه', 'شناسه دوره', 'عنوان دوره',
            'نوع هزینه', 'مبلغ', 'توضیحات', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت هزینه‌ها
        expenses = models.Expense.objects.filter(project=project).select_related('period').order_by('period__year', 'period__month_number', 'id')
        
        # نوشتن داده‌ها
        for row_num, expense in enumerate(expenses, 2):
            data = [
                expense.id,
                expense.project.id,
                expense.period.id,
                expense.period.label,
                expense.get_expense_type_display(),
                float(expense.amount),
                expense.description,
                ExcelStyleHelper.normalize_datetime(expense.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت مبلغ
                if col_num == 6:
                    cell.number_format = '#,##0.00'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class SalesSheet:
    """شیت فروش/مرجوعی‌ها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت فروش/مرجوعی‌ها"""
        ws = workbook.create_sheet(title="Sales")
        
        # هدرها
        headers = [
            'ID', 'شناسه پروژه', 'شناسه دوره', 'عنوان دوره',
            'مبلغ', 'توضیحات', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت فروش‌ها
        sales = models.Sale.objects.filter(project=project).select_related('period').order_by('period__year', 'period__month_number', 'id')
        
        # نوشتن داده‌ها
        for row_num, sale in enumerate(sales, 2):
            data = [
                sale.id,
                sale.project.id,
                sale.period.id,
                sale.period.label,
                float(sale.amount),
                sale.description,
                ExcelStyleHelper.normalize_datetime(sale.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت مبلغ
                if col_num == 5:
                    cell.number_format = '#,##0.00'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class UserProfilesSheet:
    """شیت پروفایل کاربران"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت پروفایل کاربران"""
        ws = workbook.create_sheet(title="UserProfiles")
        
        # هدرها
        headers = [
            'ID', 'شناسه کاربر', 'نام کاربری', 'نقش',
            'شماره تلفن', 'بخش', 'تاریخ ایجاد'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت پروفایل‌های کاربران
        user_profiles = models.UserProfile.objects.select_related('user').order_by('id')
        
        # نوشتن داده‌ها
        for row_num, profile in enumerate(user_profiles, 2):
            data = [
                profile.id,
                profile.user.id,
                profile.user.username,
                profile.get_role_display(),
                profile.phone or '',
                profile.department or '',
                ExcelStyleHelper.normalize_datetime(profile.created_at)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class DashboardSheet:
    """شیت داشبورد - خلاصه کل پروژه"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت داشبورد"""
        ws = workbook.create_sheet(title="Dashboard")
        
        # دریافت محاسبات
        stats = ProjectCalculations.calculate_project_statistics(project.id)
        
        if 'error' in stats:
            # در صورت خطا، یک شیت خالی با پیام خطا
            ws.cell(row=1, column=1, value="خطا در دریافت اطلاعات")
            ws.cell(row=2, column=1, value=stats['error'])
            return ws
        
        # عنوان
        title_cell = ws.cell(row=1, column=1, value=f"داشبورد پروژه: {project.name}")
        title_cell.font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        # فاصله
        row = 3
        
        # بخش اطلاعات واحدها
        section_cell = ws.cell(row=row, column=1, value="اطلاعات واحدها")
        section_cell.font = Font(name='Tahoma', size=12, bold=True)
        section_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        units_data = [
            ('تعداد واحدها', stats.get('units_statistics', {}).get('total_units', 0)),
            ('متراژ کل', f"{stats.get('units_statistics', {}).get('total_area', 0):,.2f}"),
            ('ارزش کل (تومان)', f"{stats.get('units_statistics', {}).get('total_price', 0):,.2f}"),
            ('زیربنای کل', f"{stats.get('project', {}).get('total_infrastructure', 0):,.2f}"),
        ]
        
        for label, value in units_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # بخش اطلاعات مالی
        section_cell = ws.cell(row=row, column=1, value="اطلاعات مالی")
        section_cell.font = Font(name='Tahoma', size=12, bold=True)
        section_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        financial_data = [
            ('سرمایه موجود (تومان)', f"{stats.get('transaction_statistics', {}).get('net_principal', 0):,.2f}"),
            ('سود کل (تومان)', f"{stats.get('transaction_statistics', {}).get('total_profits', 0):,.2f}"),
            ('موجودی کل (تومان)', f"{stats.get('transaction_statistics', {}).get('grand_total', 0):,.2f}"),
            ('مجموع هزینه‌ها (تومان)', f"{stats.get('expense_statistics', {}).get('total_expenses', 0):,.2f}"),
            ('مجموع فروش (تومان)', f"{stats.get('expense_statistics', {}).get('total_sales', 0):,.2f}"),
            ('مانده صندوق (تومان)', f"{stats.get('expense_statistics', {}).get('final_cost', 0):,.2f}"),
        ]
        
        for label, value in financial_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # بخش سرمایه‌گذاران
        section_cell = ws.cell(row=row, column=1, value="اطلاعات سرمایه‌گذاران")
        section_cell.font = Font(name='Tahoma', size=12, bold=True)
        section_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        investor_data = [
            ('تعداد کل سرمایه‌گذاران', stats.get('investor_statistics', {}).get('total_investors', 0)),
            ('تعداد مالکان', stats.get('investor_statistics', {}).get('owners_count', 0)),
        ]
        
        for label, value in investor_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        
        # اعمال استایل به همه سلول‌ها
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value:
                    cell.font = ExcelStyleHelper.get_default_font()
                    cell.border = ExcelStyleHelper.get_border()
        
        return ws


class ProfitMetricsSheet:
    """شیت محاسبات سود"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت محاسبات سود"""
        ws = workbook.create_sheet(title="Profit_Metrics")
        
        # دریافت محاسبات
        metrics = ProfitCalculations.calculate_profit_percentages(project.id)
        
        if 'error' in metrics:
            ws.cell(row=1, column=1, value="خطا در دریافت اطلاعات")
            ws.cell(row=2, column=1, value=metrics['error'])
            return ws
        
        # عنوان
        title_cell = ws.cell(row=1, column=1, value="محاسبات سود")
        title_cell.font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # داده‌ها
        data = [
            ('درصد سود کل', f"{metrics.get('total_profit_percentage', 0):.2f}%"),
            ('درصد سود سالانه', f"{metrics.get('annual_profit_percentage', 0):.2f}%"),
            ('درصد سود ماهانه', f"{metrics.get('monthly_profit_percentage', 0):.2f}%"),
            ('درصد سود روزانه', f"{metrics.get('daily_profit_percentage', 0):.6f}%"),
            ('دوره متوسط ساخت (ماه)', f"{metrics.get('average_construction_period', 0):.2f}"),
            ('ضریب اصلاحی', f"{metrics.get('correction_factor', 1):.10f}"),
        ]
        
        for label, value in data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # اعمال استایل
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value:
                    cell.font = ExcelStyleHelper.get_default_font()
                    cell.border = ExcelStyleHelper.get_border()
        
        return ws


class CostMetricsSheet:
    """شیت محاسبات هزینه"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت محاسبات هزینه"""
        ws = workbook.create_sheet(title="Cost_Metrics")
        
        # دریافت محاسبات
        metrics = ProjectCalculations.calculate_cost_metrics(project.id)
        
        if 'error' in metrics:
            ws.cell(row=1, column=1, value="خطا در دریافت اطلاعات")
            ws.cell(row=2, column=1, value=metrics['error'])
            return ws
        
        # عنوان
        title_cell = ws.cell(row=1, column=1, value="محاسبات هزینه")
        title_cell.font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # داده‌ها
        data = [
            ('مجموع هزینه‌ها (تومان)', f"{metrics.get('total_expenses', 0):,.2f}"),
            ('مجموع فروش (تومان)', f"{metrics.get('total_sales', 0):,.2f}"),
            ('هزینه نهایی (تومان)', f"{metrics.get('final_cost', 0):,.2f}"),
            ('ارزش کل (تومان)', f"{metrics.get('total_value', 0):,.2f}"),
            ('سود نهایی (تومان)', f"{metrics.get('final_profit_amount', 0):,.2f}"),
            ('درصد سود کل', f"{metrics.get('total_profit_percentage', 0):.2f}%"),
            ('', ''),
            ('هزینه هر متر خالص (تومان)', f"{metrics.get('net_cost_per_meter', 0):,.2f}"),
            ('هزینه هر متر ناخالص (تومان)', f"{metrics.get('gross_cost_per_meter', 0):,.2f}"),
            ('ارزش هر متر (تومان)', f"{metrics.get('value_per_meter', 0):,.2f}"),
            ('', ''),
            ('متراژ کل', f"{metrics.get('total_area', 0):,.2f}"),
            ('زیربنای کل', f"{metrics.get('total_infrastructure', 0):,.2f}"),
        ]
        
        for label, value in data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        
        # اعمال استایل
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value:
                    cell.font = ExcelStyleHelper.get_default_font()
                    cell.border = ExcelStyleHelper.get_border()
        
        return ws


class InvestorAnalysisSheet:
    """شیت تحلیل سرمایه‌گذاران"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت تحلیل سرمایه‌گذاران"""
        ws = workbook.create_sheet(title="Investor_Analysis")
        
        # هدرها
        headers = [
            'نام سرمایه‌گذار', 'آورده کل', 'برداشت کل', 'سرمایه خالص',
            'سود کل', 'موجودی کل', 'نسبت سرمایه (%)', 'نسبت سود (%)',
            'نسبت موجودی کل (%)', 'شاخص نفع',
            # اطلاعات مالکیتی
            'تعداد واحدها', 'متراژ مالکیت', 'مجموع متراژ واحدها', 'درصد مالکیت',
            'قیمت متوسط/متر', 'مجموع قیمت واحدها', 'قیمت واگذاری/متر',
            'پرداخت نهایی', 'مبلغ واقعی پرداختی'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت همه سرمایه‌گذاران
        investors = models.Investor.objects.all().order_by('last_name', 'first_name')
        
        # نوشتن داده‌ها
        for row_num, investor in enumerate(investors, 2):
            # محاسبات مستقیم برای سرمایه‌گذار
            transactions = models.Transaction.objects.filter(
                investor=investor,
                project=project
            )
            
            # محاسبه مبالغ
            total_principal = transactions.filter(
                transaction_type='principal_deposit'
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            total_withdrawal = transactions.filter(
                transaction_type='principal_withdrawal'
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            total_profit = transactions.filter(
                transaction_type='profit_accrual'
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            # محاسبات اصلی
            net_principal = float(total_principal) + float(total_withdrawal)  # withdrawal منفی است
            total_balance = net_principal + float(total_profit)
            
            # محاسبه نسبت‌ها (ساده)
            try:
                # دریافت آمار کل پروژه برای محاسبه نسبت‌ها
                project_stats = ProjectCalculations.calculate_project_statistics(project.id)
                total_project_principal = project_stats.get('transaction_statistics', {}).get('net_principal', 1)
                total_project_profits = project_stats.get('transaction_statistics', {}).get('total_profits', 1)
                total_project_balance = project_stats.get('transaction_statistics', {}).get('grand_total', 1)
                
                capital_ratio = (net_principal / total_project_principal * 100) if total_project_principal != 0 else 0
                profit_ratio = (float(total_profit) / total_project_profits * 100) if total_project_profits != 0 else 0
                total_ratio = (total_balance / total_project_balance * 100) if total_project_balance != 0 else 0
                
                # شاخص نفع
                profit_index = (profit_ratio / capital_ratio) if capital_ratio != 0 else 0
                
            except:
                capital_ratio = profit_ratio = total_ratio = profit_index = 0
            
            # محاسبه اطلاعات مالکیتی
            try:
                from construction.calculations import InvestorCalculations
                ownership_data = InvestorCalculations.calculate_investor_ownership(investor.id, project.id)
                
                units_count = ownership_data.get('units_count', 0)
                ownership_area = ownership_data.get('ownership_area', 0)
                total_units_area = ownership_data.get('total_units_area', 0)
                ownership_percentage = ownership_data.get('ownership_percentage', 0)
                average_price_per_meter = ownership_data.get('average_price_per_meter', 0)
                total_units_price = ownership_data.get('total_units_price', 0)
                transfer_price_per_meter = ownership_data.get('transfer_price_per_meter', 0)
                final_payment = ownership_data.get('final_payment', 0)
                actual_paid = ownership_data.get('actual_paid', 0)
            except:
                units_count = ownership_area = total_units_area = ownership_percentage = 0
                average_price_per_meter = total_units_price = transfer_price_per_meter = 0
                final_payment = actual_paid = 0
            
            data = [
                f"{investor.first_name} {investor.last_name}",
                float(total_principal),
                float(total_withdrawal),
                net_principal,
                float(total_profit),
                total_balance,
                capital_ratio,
                profit_ratio,
                total_ratio,
                profit_index,
                # اطلاعات مالکیتی
                units_count,
                ownership_area,
                total_units_area,
                ownership_percentage,
                average_price_per_meter,
                total_units_price,
                transfer_price_per_meter,
                final_payment,
                actual_paid,
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت اعداد
                if col_num in [2, 3, 4, 5, 6, 16, 18, 19]:  # مبالغ (اضافه شد: مجموع قیمت، پرداخت نهایی، مبلغ واقعی)
                    cell.number_format = '#,##0.00'
                elif col_num in [7, 8, 9, 14]:  # درصدها (اضافه شد: درصد مالکیت)
                    cell.number_format = '0.00'
                elif col_num == 10:  # شاخص نفع
                    cell.number_format = '0.0000'
                elif col_num in [12, 13]:  # متراژ (متراژ مالکیت، مجموع متراژ)
                    cell.number_format = '#,##0.00'
                elif col_num in [15, 17]:  # قیمت هر متر (قیمت متوسط، قیمت واگذاری)
                    cell.number_format = '#,##0.00'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class PeriodSummarySheet:
    """شیت خلاصه دوره‌ای"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت خلاصه دوره‌ای"""
        ws = workbook.create_sheet(title="Period_Summary")
        
        # هدرها
        headers = [
            'عنوان دوره', 'آورده', 'برداشت', 'سرمایه خالص', 'سود',
            'هزینه‌ها', 'فروش', 'مانده صندوق',
            'آورده تجمعی', 'برداشت تجمعی', 'سرمایه تجمعی', 'سود تجمعی',
            'هزینه تجمعی', 'فروش تجمعی', 'مانده تجمعی'
        ]
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            ExcelStyleHelper.apply_header_style(cell)
        
        # دریافت دوره‌ها
        periods = models.Period.objects.filter(project=project).order_by('year', 'month_number')
        
        # متغیرهای تجمعی
        cumulative_deposits = 0
        cumulative_withdrawals = 0
        cumulative_net_capital = 0
        cumulative_profits = 0
        cumulative_expenses = 0
        cumulative_sales = 0
        final_fund_balance = 0
        
        row_num = 2
        
        # نوشتن داده‌ها
        for period in periods:
            # محاسبه تراکنش‌های دوره
            period_transactions = models.Transaction.objects.filter(
                project=project,
                period=period
            )
            
            # آورده
            deposits = period_transactions.filter(
                transaction_type='principal_deposit'
            ).aggregate(total=Sum('amount'))['total'] or 0
            deposits = float(deposits)
            cumulative_deposits += deposits
            
            # برداشت
            withdrawals = period_transactions.filter(
                transaction_type='principal_withdrawal'
            ).aggregate(total=Sum('amount'))['total'] or 0
            withdrawals = float(withdrawals)
            cumulative_withdrawals += withdrawals
            
            # سود
            profits = period_transactions.filter(
                transaction_type='profit_accrual'
            ).aggregate(total=Sum('amount'))['total'] or 0
            profits = float(profits)
            cumulative_profits += profits
            
            # سرمایه خالص دوره (withdrawals منفی است)
            net_capital = deposits + withdrawals
            cumulative_net_capital += net_capital
            
            # هزینه‌های دوره
            expenses = models.Expense.objects.filter(
                project=project,
                period=period
            ).aggregate(total=Sum('amount'))['total'] or 0
            expenses = float(expenses)
            cumulative_expenses += expenses
            
            # فروش/مرجوعی دوره
            sales = models.Sale.objects.filter(
                project=project,
                period=period
            ).aggregate(total=Sum('amount'))['total'] or 0
            sales = float(sales)
            cumulative_sales += sales
            
            # مانده صندوق
            fund_balance = cumulative_net_capital - cumulative_expenses + cumulative_sales
            final_fund_balance = fund_balance
            
            # داده‌های ردیف
            data = [
                period.label,
                deposits,
                withdrawals,
                net_capital,
                profits,
                expenses,
                sales,
                fund_balance,
                cumulative_deposits,
                cumulative_withdrawals,
                cumulative_net_capital,
                cumulative_profits,
                cumulative_expenses,
                cumulative_sales,
                fund_balance
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                ExcelStyleHelper.apply_cell_style(cell)
                
                # فرمت مبالغ
                if col_num > 1:
                    cell.number_format = '#,##0.00'
            
            row_num += 1
        
        # افزودن ردیف جمع کل
        total_data = [
            'جمع کل',
            cumulative_deposits,
            cumulative_withdrawals,
            cumulative_net_capital,
            cumulative_profits,
            cumulative_expenses,
            cumulative_sales,
            final_fund_balance,
            '', '', '', '', '', '', ''
        ]
        
        for col_num, value in enumerate(total_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name='Tahoma', size=10, bold=True)
            cell.border = ExcelStyleHelper.get_border()
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            
            if col_num > 1 and value != '':
                cell.number_format = '#,##0.00'
        
        # تنظیم عرض ستون‌ها
        ExcelStyleHelper.auto_adjust_column_width(ws)
        
        return ws


class TransactionSummarySheet:
    """شیت خلاصه تراکنش‌ها"""
    
    @staticmethod
    def create(workbook, project):
        """ایجاد شیت خلاصه تراکنش‌ها"""
        ws = workbook.create_sheet(title="Transaction_Summary")
        
        # دریافت محاسبات
        stats = TransactionCalculations.calculate_transaction_statistics(project_id=project.id)
        
        if 'error' in stats:
            ws.cell(row=1, column=1, value="خطا در دریافت اطلاعات")
            ws.cell(row=2, column=1, value=stats['error'])
            return ws
        
        # عنوان
        title_cell = ws.cell(row=1, column=1, value="خلاصه تراکنش‌ها")
        title_cell.font = Font(name='Tahoma', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # داده‌ها
        data = [
            ('تعداد کل تراکنش‌ها', stats.get('total_transactions', 0)),
            ('', ''),
            ('مجموع آورده‌ها (تومان)', f"{stats.get('total_deposits', 0):,.2f}"),
            ('مجموع برداشت‌ها (تومان)', f"{stats.get('total_withdrawals', 0):,.2f}"),
            ('سرمایه موجود (تومان)', f"{stats.get('net_capital', 0):,.2f}"),
            ('', ''),
            ('مجموع سودها (تومان)', f"{stats.get('total_profits', 0):,.2f}"),
        ]
        
        for label, value in data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        # اعمال استایل
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value:
                    cell.font = ExcelStyleHelper.get_default_font()
                    cell.border = ExcelStyleHelper.get_border()
        
        return ws


class ExcelExportService:
    """سرویس اصلی تولید فایل Excel"""
    
    def __init__(self, project=None):
        """
        مقداردهی اولیه
        
        Args:
            project: پروژه مورد نظر (اگر None باشد، پروژه فعال استفاده می‌شود)
        """
        self.project = project or models.Project.get_active_project()
        self.workbook = Workbook()
    
    def generate_excel(self):
        """
        تولید فایل Excel کامل با تمام شیت‌ها
        
        Returns:
            Workbook: فایل Excel آماده
        """
        if not self.project:
            raise ValueError('هیچ پروژه‌ای برای export یافت نشد')
        
        # حذف شیت پیش‌فرض
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # ایجاد شیت فهرست محتوا (در ابتدا)
        TableOfContentsSheet.create(self.workbook, self.project)
        
        # ایجاد شیت‌های پایه
        self.create_base_sheets()
        
        # ایجاد شیت‌های محاسباتی
        self.create_calculation_sheets()
        
        return self.workbook
    
    def create_base_sheets(self):
        """ایجاد شیت‌های داده پایه"""
        # ایجاد شیت‌ها
        ProjectSheet.create(self.workbook, self.project)
        UnitsSheet.create(self.workbook, self.project)
        InvestorsSheet.create(self.workbook, self.project)
        PeriodsSheet.create(self.workbook, self.project)
        InterestRatesSheet.create(self.workbook, self.project)
        TransactionsSheet.create(self.workbook, self.project)
        ExpensesSheet.create(self.workbook, self.project)
        SalesSheet.create(self.workbook, self.project)
        UserProfilesSheet.create(self.workbook, self.project)
        
        # اضافه کردن فریز و فیلتر به همه شیت‌های داده پایه
        base_sheets = ['Units', 'Investors', 'Periods', 'InterestRates', 
                      'Transactions', 'Expenses', 'Sales', 'UserProfiles']
        for sheet_name in base_sheets:
            if sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
                ExcelStyleHelper.freeze_header_row(ws)
                ExcelStyleHelper.add_auto_filter(ws)
    
    def create_calculation_sheets(self):
        """ایجاد شیت‌های محاسباتی"""
        DashboardSheet.create(self.workbook, self.project)
        ProfitMetricsSheet.create(self.workbook, self.project)
        CostMetricsSheet.create(self.workbook, self.project)
        InvestorAnalysisSheet.create(self.workbook, self.project)
        PeriodSummarySheet.create(self.workbook, self.project)
        TransactionSummarySheet.create(self.workbook, self.project)

